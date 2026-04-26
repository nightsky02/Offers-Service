from app.excel.models import OfferDataModel, WrongOfferModel, OfferParsingResult
from app.excel.exceptions import UniqueOfferException, FileStructureError
from openpyxl import load_workbook
from pydantic import ValidationError


def _format_exception(ex: ValidationError) -> str:
    first_error = ex.errors()[0]

    return f"*{first_error["loc"][0]}* -  {first_error['msg']}"


def contains_offer_by_id(offer_id: int, offers: list[OfferDataModel]) -> bool:
    for offer in offers:
        if offer_id == offer.offer_id:
            return True
    return False


def is_end_of_file(row: list) -> bool:
    empty_columns = 0
    for col_value in row:
        if col_value is None:
            empty_columns += 1
        
        if isinstance(col_value, str) and not len(col_value):
            empty_columns += 1

    return empty_columns >= len(row)


def parse_file(file_object) -> OfferParsingResult:

    workbook = load_workbook(file_object)
    sheet = workbook.active

    offers = []
    wrong_offers = [] # for invalid offers
     
    # skip the first row, because there are headers of the columns by default
    for row in sheet.iter_rows(min_row=2):
        table_values = [c.value for c in row]

        if len(table_values) != 5: # 5 cells, check the rules of excel file
            raise FileStructureError("The wrong count of the cells")

        offer_id = table_values[0]

        # if the row is fully empty, then consider it as the end of the document
        if is_end_of_file(table_values):
            break
        elif offer_id is None: 
            # if only offer id column is empty, then it's the wrong file structure
            raise FileStructureError("No offer id for row")
        if contains_offer_by_id(offer_id, offers):
            raise UniqueOfferException(offer_id, table_values[1])
        
        try:
            offers.append(OfferDataModel(
                offer_id=offer_id,
                name=table_values[1],
                price=table_values[2],
                quantity=table_values[3],
                avaivable=table_values[4]
            ))
        except ValidationError as e:
            wrong_offers.append(WrongOfferModel(
                offer_id=table_values[0],
                err_msg=_format_exception(e)))
            
    return OfferParsingResult(correct_offers=offers, wrong_offers=wrong_offers)
